import os
import shutil
from flask import Flask, render_template, request, flash, redirect, url_for, send_file
from openpyxl import load_workbook

app = Flask(__name__)
app.config['SECRET_KEY'] = 'LBG' #clé secrete à ne pas donner en prod

# CONSTANTES
IMG_PATH = '/PROD_LBG_FACTURE_AUTO/templates/img/'
REPOSITORY_FACTURES_PATH = '/PROD_LBG_FACTURE_AUTO/factures/'

# Fonctions utiles 
def SUPPR_REP(rep_path): #supprime les fichiers d'un répertoire
     if not os.path.isdir(rep_path):
        return 'Dossier inexistant'
     for fichier in os.listdir(rep_path):
        chemin_fichier = os.path.join(rep_path, fichier)
        
        
        if os.path.isfile(chemin_fichier): #Vérification si le chemin est un fichier
            os.remove(chemin_fichier) #Supprimer le fichier


# CLASSES
# Représentation du pdf à partir du html
class Pdf():
    def render_pdf(self,repository, name, html):
        from xhtml2pdf import pisa
        with open(repository+name+".pdf","wb") as pdf:
            ret = pisa.CreatePDF(html, pdf)
            if ret.err:
                flash("Erreur lors de la génération des pdf", category='error')
                return ""
        return pdf

# Réprésentation des données dans le fichier Excel  
class DataExcel:
    # Fonction qui retourne la feuille contenant les données
    def get_data_from_excel(file:str): 
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        return sheet

#Représente une facture
class Facture:
    def __init__(self, facture, client) -> None:
        self.facture = facture
        self.client = client
        self.lst_lignes = []
        self.total = 0
    
    def getLignes(self):
        return self.lst_lignes
    
    def addLignes(self, ligne_fct):
        self.lst_lignes.append(ligne_fct)

    def calculTotal(self):
        total = 0
        for ligne in self.lst_lignes:
            total = (ligne.qte*ligne.prix)+total
        return total

#Représente les lignes d'une facture
class LigneFacture:
    def __init__(self, article, qte, prix) -> None:
        self.article = article
        self.qte = qte
        self.prix = prix

    def Print_toString(self): # utile pour les tests
        print("article : ", self.article)
        print("quantité : ", self.qte)
        print("prix : ", self.prix) 


#Représente une famille
class Famille:
    def __init__(self, nom_famille) -> None: #ajouté une facture à la famille ? (pour gérer date etc ?)
        self.nom_famille = nom_famille
        self.lst_membre_fam = []
    
    def getMembreFam(self):
        return self.lst_membre_fam
    
    def addMembreFam(self, membre_fam):
        self.lst_membre_fam.append(membre_fam)
    
    def calculTotal(self):
        total = 0
        for membre in self.lst_membre_fam:
            total = (membre.heures*membre.tarif)+total
        return total

#Représente un membre d'une famille
class Membre:
    def __init__(self, prenom_membre, heures) -> None:
        self.prenom_membre = prenom_membre
        self.heures = heures
        self.tarif = 1 #le tarif d'un entrainement est de 1 euros => A CHANGER OU A METTRE DANS LE FICHIER EXCEL ?



# VUES
@app.route('/', methods=['GET','POST'])
def home():
    return render_template('LBG_home.html')

# Création des factures pour les entrainements
@app.route('/entrainement', methods=['GET','POST'])
def entrainement():
    if request.method == 'POST':
        # Récupération du fichier Excel
        file = request.files['file_entrainement']
        # On vide le répertoire de dépot
        SUPPR_REP(REPOSITORY_FACTURES_PATH)
        
        if file.filename == '':
            flash('Aucun fichier sélectionné')
            return redirect(url_for('home'))

        data = DataExcel.get_data_from_excel(file)
        
        # Création des familles dans un dictionnaire
        lst_famille = {}
        for num_row in range(7, data.max_row-1):
            nom = data.cell(num_row,2).value
            prenom = data.cell(num_row,3).value
            heures = data.cell(num_row,40).value

            famille = None
            if heures != 0:
                gymnaste = Membre(prenom, heures)
                if nom in lst_famille:
                    famille = lst_famille[nom]
                else:
                    famille = Famille(nom)
                famille.addMembreFam(gymnaste)
                lst_famille[nom] = famille
        
        # Création des factures par famille
        nb_fac = 0
        for nom_fam in lst_famille:
            famille = lst_famille[nom_fam]
            template_facture = render_template('LBG_template_entrainement.html',
                                    nom_famille = famille.nom_famille,
                                    membres_famille = famille.getMembreFam(),
                                    total = famille.calculTotal(),
                                    img_path=IMG_PATH)
            pdf = Pdf()
            pdf.render_pdf(REPOSITORY_FACTURES_PATH,'Facture_'+famille.nom_famille,template_facture)
            nb_fac = nb_fac+1
        flash("Création de "+str(nb_fac)+" factures avec succès ! ", category="success")
    return redirect(url_for('home'))

# Téléchargement des factures dans un fichier zip
@app.route('/zip_factures', methods=['GET','POST'])
def zip_factures():
    zip_filename = 'factures.zip'  # Nom du fichier ZIP

    # Vérifiez si le chemin du répertoire est valide
    if not os.path.isdir(REPOSITORY_FACTURES_PATH):
        return 'Répertoire invalide'

    # Créez le fichier ZIP
    shutil.make_archive(zip_filename[:-4], 'zip', REPOSITORY_FACTURES_PATH)

    # Envoyez le fichier ZIP au navigateur de l'utilisateur
    return send_file(zip_filename, as_attachment=True)




if __name__ == '__main__':
    app.run(debug=True)